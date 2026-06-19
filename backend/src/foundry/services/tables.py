import json
import re
import threading
from pathlib import Path
from typing import Any

import duckdb
import sqlglot
from openpyxl import load_workbook
from sqlglot import exp

from foundry.core.errors import ValidationError

IDENTIFIER_PATTERN = re.compile(r"[^a-zA-Z0-9_]")
FORBIDDEN_SQL = (
    exp.Insert,
    exp.Update,
    exp.Delete,
    exp.Create,
    exp.Drop,
    exp.Alter,
    exp.Command,
    exp.Merge,
)


def safe_identifier(value: str) -> str:
    normalized = IDENTIFIER_PATTERN.sub("_", value).strip("_").lower()
    return (normalized or "table")[:48]


class TableStore:
    def __init__(self, database_path: Path) -> None:
        database_path.parent.mkdir(parents=True, exist_ok=True)
        self.connection = duckdb.connect(str(database_path))
        self.lock = threading.Lock()

    def close(self) -> None:
        with self.lock:
            self.connection.close()

    def import_file(self, path: Path, table_name: str) -> None:
        with self.lock:
            quoted_table = f'"{table_name}"'
            self.connection.execute(f"DROP TABLE IF EXISTS {quoted_table}")
            if path.suffix.lower() == ".csv":
                self.connection.execute(
                    f"CREATE TABLE {quoted_table} AS SELECT * FROM read_csv_auto(?)", [str(path)]
                )
            elif path.suffix.lower() in {".xlsx", ".xlsm"}:
                self._import_excel(path, quoted_table)
            else:
                raise ValidationError("TAG supports CSV and XLSX files")

    def _import_excel(self, path: Path, quoted_table: str) -> None:
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        try:
            raw_headers = next(rows)
        except StopIteration as exc:
            raise ValidationError("Excel workbook is empty") from exc
        headers = [
            safe_identifier(str(value or f"column_{index + 1}"))
            for index, value in enumerate(raw_headers)
        ]
        unique_headers: list[str] = []
        for index, header in enumerate(headers):
            candidate = header
            while candidate in unique_headers:
                candidate = f"{header}_{index + 1}"
            unique_headers.append(candidate)
        columns = ", ".join(f'"{header}" VARCHAR' for header in unique_headers)
        self.connection.execute(f"CREATE TABLE {quoted_table} ({columns})")
        placeholders = ", ".join("?" for _ in unique_headers)
        values = [tuple(None if value is None else str(value) for value in row) for row in rows]
        if values:
            self.connection.executemany(
                f"INSERT INTO {quoted_table} VALUES ({placeholders})", values
            )

    def schema(self, table_name: str) -> list[dict[str, str]]:
        with self.lock:
            rows = self.connection.execute(f'PRAGMA table_info("{table_name}")').fetchall()
        return [{"name": str(row[1]), "type": str(row[2])} for row in rows]

    def sample(self, table_name: str, limit: int = 3) -> list[dict[str, Any]]:
        with self.lock:
            cursor = self.connection.execute(f'SELECT * FROM "{table_name}" LIMIT ?', [limit])
            columns = [description[0] for description in cursor.description]
            rows = cursor.fetchall()
        return [dict(zip(columns, row, strict=True)) for row in rows]

    def catalog_text(self, table_name: str) -> str:
        schema = self.schema(table_name)
        sample = self.sample(table_name)
        return (
            f"Table: {table_name}\n"
            f"Columns: {json.dumps(schema, ensure_ascii=False)}\n"
            f"Sample rows: {json.dumps(sample, ensure_ascii=False, default=str)}"
        )

    def execute_safe(self, sql: str, max_rows: int = 100) -> dict[str, Any]:
        cleaned = sql.strip().rstrip(";")
        try:
            statements = sqlglot.parse(cleaned, read="duckdb")
        except sqlglot.errors.ParseError as exc:
            raise ValidationError("Generated SQL is not valid") from exc
        if len(statements) != 1 or statements[0].find(exp.Select) is None:
            raise ValidationError("Only one SELECT query is allowed")
        if any(statements[0].find(node_type) is not None for node_type in FORBIDDEN_SQL):
            raise ValidationError("SQL contains a forbidden operation")

        bounded_sql = f"SELECT * FROM ({cleaned}) AS foundry_query LIMIT {max_rows}"
        try:
            with self.lock:
                cursor = self.connection.execute(bounded_sql)
                columns = [description[0] for description in cursor.description]
                rows = cursor.fetchall()
        except duckdb.Error as exc:
            raise ValidationError(f"TAG query failed: {exc}") from exc
        return {
            "sql": cleaned,
            "columns": columns,
            "rows": [list(row) for row in rows],
        }
